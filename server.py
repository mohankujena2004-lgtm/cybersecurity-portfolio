# server.py
from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl, os, subprocess, shlex
import logging
from pathlib import Path

app = Flask(__name__)
CORS(app)  # allow frontend access (you can restrict origins in production)

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")

BASE_DIR = Path(__file__).resolve().parent
FILE_PATH = BASE_DIR / "user.xlsx"

# read API key from environment
API_KEY = os.environ.get("API_KEY", None)
if not API_KEY:
    logging.warning("API_KEY not set. Set environment variable API_KEY in deployment!")

SCAN_MAP = {
    "tcp": "-sT",
    "syn": "-sS",
    "fin": "-sF",
    "null": "-sN",
    "xmas": "-sX",
    "ack": "-sA",
    "udp": "-sU",
    "version": "-sV"
}

def require_api_key(req):
    if not API_KEY:
        return True  # if not set, allow (useful for local dev) — but in prod always set API_KEY
    key = req.headers.get("x-api-key") or req.args.get("api_key")
    return key == API_KEY

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"}), 200

# ---------- REGISTER ----------
@app.route('/register', methods=['POST'])
def register_user():
    try:
        if not require_api_key(request):
            return jsonify({"message": "Unauthorized"}), 401

        data = request.get_json(force=True)
        email = data.get("email")
        phone = data.get("phone")
        password = data.get("password")

        if not email or not phone or not password:
            return jsonify({"message": "Missing fields"}), 400

        # create Excel if missing
        if not FILE_PATH.exists():
            wb = openpyxl.Workbook()
            sh = wb.active
            sh.title = "Users"
            sh.append(["Email", "Phone", "Password"])
            wb.save(FILE_PATH)

        wb = openpyxl.load_workbook(FILE_PATH)
        sh = wb.active

        for row in sh.iter_rows(min_row=2, values_only=True):
            if email == row[0] or phone == row[1]:
                wb.close()
                return jsonify({"message": "User already exists"}), 409

        sh.append([email, phone, password])
        wb.save(FILE_PATH)
        wb.close()

        logging.info(f"Registered user: {email}")
        return jsonify({"message": "User registered successfully"}), 200

    except Exception as e:
        logging.exception("Error in register_user")
        return jsonify({"message": f"Server error: {str(e)}"}), 500

# ---------- LOGIN ----------
@app.route('/login', methods=['POST'])
def login_user():
    try:
        if not require_api_key(request):
            return jsonify({"message": "Unauthorized"}), 401

        data = request.get_json(force=True)
        email_or_phone = data.get("emailOrPhone")
        password = data.get("password")

        if not email_or_phone or not password:
            return jsonify({"message": "Missing fields"}), 400

        if not FILE_PATH.exists():
            return jsonify({"message": "No users registered yet!"}), 404

        wb = openpyxl.load_workbook(FILE_PATH)
        sh = wb.active

        for row in sh.iter_rows(min_row=2, values_only=True):
            email, phone, stored_password = row
            if (email_or_phone == email or email_or_phone == phone) and password == stored_password:
                wb.close()
                logging.info(f"Login success for: {email_or_phone}")
                return jsonify({"message": "Login successful", "user": {"email": email, "phone": phone}}), 200

        wb.close()
        logging.warning(f"Invalid login attempt: {email_or_phone}")
        return jsonify({"message": "Invalid credentials"}), 401

    except Exception as e:
        logging.exception("Error in login_user")
        return jsonify({"message": f"Server error: {str(e)}"}), 500

# ---------- SCAN ----------
@app.route('/scan', methods=['POST'])
def run_scan():
    try:
        if not require_api_key(request):
            return jsonify({"message": "Unauthorized"}), 401

        data = request.get_json(force=True)
        target = data.get('target')
        stype = data.get('type', 'tcp')
        ports = data.get('ports')  # optional

        logging.info(f"Scan request received: target={target}, type={stype}, ports={ports}")

        if not target:
            return jsonify({"error": "target required"}), 400
        if stype not in SCAN_MAP:
            return jsonify({"error": "unsupported scan type"}), 400

        flag = SCAN_MAP[stype]
        port_arg = f"-p {ports}" if ports else ""
        cmd = f"nmap {flag} {port_arg} -Pn -T4 {target}".strip()
        logging.info("Running command: %s", cmd)

        try:
            proc = subprocess.run(shlex.split(cmd), capture_output=True, text=True, timeout=180)
        except FileNotFoundError:
            logging.error("nmap not found on server")
            return jsonify({"error": "nmap not found on server. Install nmap."}), 500
        except subprocess.TimeoutExpired:
            logging.error("nmap scan timed out")
            return jsonify({"error": "scan timed out"}), 500

        output = proc.stdout.strip() or proc.stderr.strip() or ""
        return jsonify({"output": output}), 200

    except Exception as e:
        logging.exception("Error in run_scan")
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    logging.info("Starting Flask server on 0.0.0.0:5000")
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
